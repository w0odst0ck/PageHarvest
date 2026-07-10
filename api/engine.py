"""
PageHarvest API — 路由层

process_upload() 为唯一入口，自动识别页面类型并路由。
详情页逻辑留在此模块（尚未冻结）。
搜索页逻辑移到 api/search.py（已冻结，不可修改）。
"""

import os
import sys
import io
import re
import json
import uuid
import shutil
import zipfile
import tempfile
import subprocess
import logging
from pathlib import Path
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)


# ── 路径 ──
HERE = Path(__file__).resolve().parent
ROOT = HERE.parent          # 项目根


# ═══════════════════════════════════════════════════════════════
#  作业管理（每次请求一个 job_id，输出到独立临时目录）
# ═══════════════════════════════════════════════════════════════

class Job:
    """一次 API 调用对应一个作业"""

    def __init__(self):
        self.job_id = uuid.uuid4().hex[:12]
        self.work_dir = Path(tempfile.mkdtemp(prefix=f"ph_{self.job_id}_"))
        self.extract_dir = self.work_dir / "input"
        self.output_dir = self.work_dir / "output"

    def extract_zip(self, zip_bytes: bytes) -> list[Path]:
        """解压上传的 ZIP 到 input/ 目录，返回所有提取的文件"""
        self.extract_dir.mkdir(parents=True, exist_ok=True)
        zip_path = self.work_dir / "upload.zip"
        zip_path.write_bytes(zip_bytes)

        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(self.extract_dir)

        return sorted(self.extract_dir.rglob("*")) if any(self.extract_dir.iterdir()) else []

    def html_files(self) -> list[Path]:
        return sorted(self.extract_dir.rglob("*.html"))

    def xlsx_files(self) -> list[Path]:
        return sorted(self.extract_dir.rglob("*.xlsx"))

    def detect_platform(self, html: str) -> str:
        """从 HTML 内容检测平台"""
        if "detail.1688.com" in html or ("1688.com" in html and "cbu01.alicdn.com" in html):
            return "1688"
        if "zkh.com" in html:
            return "震坤行"
        if "item.jd.com" in html or ('class="attrs"' in html and 'class="top-name"' in html):
            return "京东"
        if "s.1688.com" in html or "1688.com" in html:
            return "1688"
        if "search.jd.com" in html or "360buyimg.com" in html:
            return "京东"
        if "zkh.com/search" in html:
            return "震坤行"
        return "未知"

    def detect_page_type(self, html: str) -> str:
        """检测页面类型：search / detail / unknown

        详情页签名优先（更具体，不易被导航栏干扰），
        仅当无详情页特征时才检查搜索页签名。
        """
        _detail_signals = [
            "detail.1688.com/offer/",  # 1688 详情页
            "module-od-product-attributes",  # 1688 详情页
            'class="attrs"',           # 京东新版详情页
            'class="highlight-attrs"', # 京东新版高亮属性
            'class="top-name"',        # 京东新版店铺名
            "sku-number",              # 震坤行详情页
            "gallery-slick-box",       # 震坤行详情页
        ]
        if any(sig in html for sig in _detail_signals):
            return "detail"

        _search_signals = [
            "zkh.com/search",          # 震坤行搜索页
            "search.jd.com",           # 京东搜索页
            "s.1688.com",              # 1688 搜索页
            "offer_search.htm",        # 1688 搜索页
            "goods-item-wrap-new",     # 震坤行搜索页
        ]
        s = html.lower()
        if any(sig in s for sig in _search_signals):
            return "search"

        return "unknown"

    def cleanup(self):
        """清理作业临时目录"""
        shutil.rmtree(self.work_dir, ignore_errors=True)


# ═══════════════════════════════════════════════════════════════
#  API：详情页解析
# ═══════════════════════════════════════════════════════════════

@dataclass
class DetailResult:
    xlsx_bytes: bytes = field(default_factory=bytes)
    csv_content: str = ""
    product_count: int = 0
    error: str = ""


def run_detail_pipeline(html_files: list[Path], job: Job) -> DetailResult:
    """详情页解析 → core.detail_parser（流式输出 Excel）"""
    result = DetailResult()

    if not html_files:
        result.error = "未找到 HTML 文件"
        return result

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        result.error = "需要 openpyxl: pip install openpyxl"
        return result

    wb = Workbook()
    ws = wb.active
    ws.title = "商品概览"

    headers = ["商品ID", "平台", "标题", "品牌", "型号", "最低价", "最高价",
               "属性数", "SKU数", "主图数"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    csv_rows = []

    for html_path in html_files:
        html = html_path.read_text(encoding="utf-8", errors="replace")
        platform = job.detect_platform(html)

        try:
            proc = subprocess.run(
                [sys.executable, "-m", "core.detail_parser", str(html_path), "--json"],
                capture_output=True, text=True, timeout=60, cwd=str(ROOT),
            )
            if proc.returncode != 0:
                ws.append([html_path.name, platform, "", "", "", "", "", "", "", "", "解析失败"])
                continue

            detail = json.loads(proc.stdout)

        except (subprocess.TimeoutExpired, json.JSONDecodeError, Exception) as e:
            ws.append([html_path.name, platform, "", "", "", "", "", "", "", "", str(e)[:60]])
            continue

        pid = detail.get("product_id", "")
        title = detail.get("title", "")
        brand = detail.get("brand", "")
        spec = detail.get("spec", "")
        pmin = detail.get("price_min", 0)
        pmax = detail.get("price_max", 0)
        attrs = detail.get("attributes", {})
        sku_count = detail.get("sku_count", 0)
        images = detail.get("main_images", [])

        ws.append([pid, platform, title[:60], brand, spec, pmin, pmax,
                   len(attrs), sku_count, len(images)])

        csv_rows.append({
            "file": html_path.name, "platform": platform, "product_id": pid,
            "title": title, "brand": brand, "spec": spec,
            "price_min": pmin, "price_max": pmax,
            "attributes": len(attrs), "sku_count": sku_count,
            "image_count": len(images), "status": "OK",
        })

        safe_name = re.sub(r'[\\/*?\[\]:]', '_', f"{title[:30] or pid}")
        sheet = wb.create_sheet(title=safe_name[:31])

        info_rows = [
            ("商品ID", pid), ("标题", title), ("品牌", brand),
            ("型号", spec), ("最低价", pmin), ("最高价", pmax),
        ]
        for kv in info_rows:
            sheet.append(kv)

        sheet.append([])
        sheet.append(["属性", "值"])
        if attrs:
            for k, v in attrs.items():
                sheet.append([k, v])

        if sku_count > 0:
            sheet.append([])
            sheet.append(["SKU 规格", "价格"])
            for sku in detail.get("sku_matrix", []):
                if isinstance(sku, dict):
                    sheet.append([sku.get("spec", ""), sku.get("price", "")])

        if images:
            sheet.append([])
            sheet.append(["主图链接"])
            for img_url in images[:20]:
                sheet.append([img_url])

    for col in ws.columns:
        try:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        except Exception:
            pass

    if csv_rows:
        import pandas as pd
        buf = io.StringIO()
        pd.DataFrame(csv_rows).to_csv(buf, index=False, encoding="utf-8-sig")
        result.csv_content = buf.getvalue()

    result.product_count = len(csv_rows)

    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_buf.seek(0)
    result.xlsx_bytes = xlsx_buf.read()

    return result


# ═══════════════════════════════════════════════════════════════
#  入口：上传 ZIP → 自动识别 → 路由
# ═══════════════════════════════════════════════════════════════

@dataclass
class ApiResult:
    success: bool = False
    page_type: str = ""       # search / detail
    platform: str = ""
    product_count: int = 0
    # 搜索页结果
    csv_content: str = ""
    txt_content: str = ""
    # 详情页结果
    xlsx_bytes: bytes = field(default_factory=bytes)
    detail_csv: str = ""
    error: str = ""


def process_upload(zip_bytes: bytes) -> ApiResult:
    """上传 ZIP → 自动识别 → 路由 → 返回结果"""
    # 搜索页处理逻辑已冻结在 api/search.py，不可修改
    from api.search import run_search_pipeline

    result = ApiResult()
    job = Job()

    try:
        all_files = job.extract_zip(zip_bytes)
        if not all_files:
            result.error = "ZIP 为空"
            return result

        html_files = job.html_files()
        xlsx_files = job.xlsx_files()

        if html_files:
            with open(html_files[0], "r", encoding="utf-8", errors="replace") as f:
                sample = f.read()
            result.platform = job.detect_platform(sample)
            result.page_type = job.detect_page_type(sample)
        elif xlsx_files:
            from api.search import detect_platform_from_name
            result.platform = detect_platform_from_name(xlsx_files[0].name)
            result.page_type = "search"
        else:
            result.error = "ZIP 中未找到 HTML 或 XLSX 文件"
            return result

        if result.page_type == "search":
            sr = run_search_pipeline(job)
            if sr.error:
                result.error = sr.error
                return result
            result.success = True
            result.csv_content = sr.csv_content
            result.txt_content = sr.txt_content
            result.product_count = sr.product_count

        elif result.page_type == "detail":
            dr = run_detail_pipeline(html_files, job)
            if dr.error:
                result.error = dr.error
                return result
            result.success = True
            result.xlsx_bytes = dr.xlsx_bytes
            result.detail_csv = dr.csv_content
            result.product_count = dr.product_count

        else:
            result.error = "无法识别页面类型（搜索页/详情页）"
            return result

        return result

    except Exception as e:
        result.error = str(e)
        return result
    finally:
        job.cleanup()
