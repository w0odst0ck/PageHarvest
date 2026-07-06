#!/usr/bin/env python3
"""
选品清单 → 详情页批量下载清单 (跨平台)
==========================================

从各平台选品输出中提取 Top 商品，生成 Excel 下载清单。
用户打开 Excel → 点击链接 → 浏览器保存 HTML → 批量解析。

用法:
    # ZKH 全品类 Top 10
    python -m selection.download_list --platform zkh --all --top 10

    # 1688 Top 10（从已有 top_products_urls.csv）
    python -m selection.download_list --platform 1688 --top 10

    # 指定品类 / 策略
    python -m selection.download_list --platform zkh --category 室内灯具 --top 20
    python -m selection.download_list --platform zkh --all --strategy "🔥 必上"
"""

import os
import re
import sys
import csv
import argparse
from pathlib import Path
from dataclasses import dataclass, field

_THIS_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = _THIS_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))


# ═══════════════════════════════════════════════════════════════
#  平台配置
# ═══════════════════════════════════════════════════════════════

@dataclass
class PlatformConfig:
    """平台配置：数据源路径、URL 模式、字段映射"""
    name: str                              # 内部名: "zkh" / "1688"
    label: str                             # 显示名: "震坤行" / "1688"
    data_dir: str                          # 数据根目录
    output_dir: str                        # 下载清单输出目录
    default_listing_subdir: str = ""       # 默认选品输出子目录
    additional_sources: list = field(default_factory=list)  # 额外 CSV 数据源

    # ── URL 处理 ──
    url_clean_regex: str = r'(https?://[^?]+\.html)'
    id_regex: str = r'/item/([^./?]+)'
    detail_url_template: str = ""          # 空则使用原始 URL

    # ── 选品 CSV 扫描 ──
    listing_pattern: str = "00-选品推荐合集.csv"
    strategy_files: list = field(default_factory=lambda: [
        "🔥 必上.csv", "👍 推荐.csv", "💡 暗马.csv", "📌 关注.csv"
    ])

    # ── 字段映射 ──
    field_roundtrip: dict = field(default_factory=lambda: {})
    """{ 输出列名: csv列名 }，空=同名"""


# ── 预设平台 ──

PLATFORMS = {
    "zkh": PlatformConfig(
        name="zkh",
        label="震坤行",
        data_dir=str(PROJECT_ROOT / "data" / "ZKH" / "震坤行"),
        output_dir=str(PROJECT_ROOT / "data" / "ZKH" / "震坤行" / "download_list"),
        default_listing_subdir="上架清单",
        url_clean_regex=r'(https?://[^?]+\.html)',
        id_regex=r'/item/([^./?]+)',
    ),
    "1688": PlatformConfig(
        name="1688",
        label="1688",
        data_dir=str(PROJECT_ROOT / "data" / "1688"),
        output_dir=str(PROJECT_ROOT / "data" / "1688" / "download_list"),
        default_listing_subdir="",
        additional_sources=["top_products_urls.csv"],
        url_clean_regex=r'(https?://detail\.1688\.com/offer/[^?]+)',
        id_regex=r'/offer/(\d+)',
        listing_pattern="*.csv",
        strategy_files=[],
    ),
}


# ═══════════════════════════════════════════════════════════════
#  URL / ID 处理
# ═══════════════════════════════════════════════════════════════

def clean_url(url: str, cfg: PlatformConfig) -> str:
    """清洗 URL，去掉追踪参数。"""
    if not url:
        return ""
    m = re.match(cfg.url_clean_regex, url)
    return m.group(1) if m else url


def extract_product_id(url: str, cfg: PlatformConfig) -> str:
    """从 URL 提取商品 ID"""
    m = re.search(cfg.id_regex, url)
    return m.group(1) if m else ""


# ═══════════════════════════════════════════════════════════════
#  读取选品数据 (ZKH 风格：目录下有品类子文件夹)
# ═══════════════════════════════════════════════════════════════

def read_listing_csv(
    csv_path: str,
    strategy_filter: str = "",
    cfg: PlatformConfig | None = None,
) -> list[dict]:
    """读取一张选品 CSV，返回统一格式列表。"""
    rows = []
    field_map = cfg.field_roundtrip if cfg else {}
    try:
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                strategy = row.get("策略", "").strip()
                title = row.get("标题", "").strip()
                url = row.get("链接", "").strip()
                brand = row.get("品牌", "").strip()
                price_str = row.get("价格", "0").strip()
                rank_str = row.get("排名", "").strip()
                model = row.get("型号", "").strip()

                if not title and not url:
                    continue

                # 无策略列 → 从文件名推断
                if not strategy:
                    fname = os.path.basename(csv_path).replace(".csv", "")
                    if fname in ("🔥 必上", "👍 推荐", "💡 暗马", "📌 关注", "🔹 补充"):
                        strategy = fname

                if strategy_filter and strategy != strategy_filter:
                    continue

                try:
                    price = float(price_str) if price_str else 0.0
                except ValueError:
                    price = 0.0
                try:
                    rank = int(rank_str) if rank_str else 0
                except ValueError:
                    rank = 0

                clean = clean_url(url, cfg) if cfg else url
                pid = extract_product_id(clean, cfg) if cfg else ""

                rows.append({
                    "品类": "",
                    "策略": strategy,
                    "排名": rank,
                    "品牌": brand,
                    "标题": title,
                    "价格": price,
                    "型号": model,
                    "商品ID": pid,
                    "详情页链接": clean,
                    "状态": "",
                })
    except Exception as e:
        print(f"  ⚠ 读取失败 {csv_path}: {e}", file=sys.stderr)
    return rows


def scan_zkh_listing(
    listing_dir: str,
    top_n: int,
    strategy_filter: str,
    cfg: PlatformConfig,
) -> list[dict]:
    """扫描 ZKH 风格选品目录（子文件夹=品类）"""
    subdirs = sorted([
        d for d in os.listdir(listing_dir)
        if os.path.isdir(os.path.join(listing_dir, d))
        and not d.startswith("_") and not d.startswith(".")
    ])

    all_items = []
    for cat in subdirs:
        cat_dir = os.path.join(listing_dir, cat)
        cat_items = []

        # 优先合集
        coll = os.path.join(cat_dir, "00-选品推荐合集.csv")
        if os.path.exists(coll):
            rows = read_listing_csv(coll, strategy_filter, cfg)
            for r in rows:
                r["品类"] = cat
            cat_items.extend(rows)

        # 各策略 CSV 补
        if not cat_items:
            for sname in cfg.strategy_files:
                scsv = os.path.join(cat_dir, sname)
                if os.path.exists(scsv):
                    rows = read_listing_csv(scsv, strategy_filter, cfg)
                    for r in rows:
                        r["品类"] = cat
                    cat_items.extend(rows)

        # 去重
        seen = set()
        uniq = []
        for item in cat_items:
            pid = item["商品ID"]
            if pid and pid in seen:
                continue
            if pid:
                seen.add(pid)
            uniq.append(item)

        uniq.sort(key=lambda x: x["排名"] if x["排名"] > 0 else 99999)
        cat_items = uniq[:top_n]
        all_items.extend(cat_items)
        print(f"  {cat:<8} {len(uniq):>3} → {len(cat_items)} 个")

    return all_items


# ═══════════════════════════════════════════════════════════════
#  读取选品数据 (1688 风格：平铺 CSV)
# ═══════════════════════════════════════════════════════════════

def scan_1688_listing(
    listing_dir: str,
    top_n: int,
    strategy_filter: str,
    cfg: PlatformConfig,
) -> list[dict]:
    """扫描 1688 风格的数据（已有 CSV 清单）"""
    all_items = []
    seen_ids = set()

    # 读取额外源
    for src in cfg.additional_sources:
        src_path = os.path.join(cfg.data_dir, src)
        if not os.path.exists(src_path):
            continue

        with open(src_path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                url = (row.get("detail_url") or "").strip()
                title = (row.get("title") or "").strip()
                shop = (row.get("shop_name") or "").strip()

                if not url:
                    continue

                clean = clean_url(url, cfg)
                pid = extract_product_id(clean, cfg)

                if pid and pid in seen_ids:
                    continue
                if pid:
                    seen_ids.add(pid)

                all_items.append({
                    "品类": "投光灯",
                    "策略": "🔝 代表品",
                    "排名": len(all_items) + 1,
                    "品牌": shop,
                    "标题": title or shop,
                    "价格": 0.0,
                    "型号": pid,
                    "商品ID": pid,
                    "详情页链接": clean,
                    "状态": "",
                })

        print(f"  {src:<35} → {len(all_items)} 个")

    # 也扫描同目录下的 CSV（如 analysis_report 等）
    for fname in sorted(os.listdir(listing_dir or cfg.data_dir)):
        if not fname.endswith(".csv"):
            continue
        fpath = os.path.join(listing_dir or cfg.data_dir, fname)
        with open(fpath, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                url = (row.get("detail_url") or row.get("链接") or "").strip()
                title = (row.get("title") or row.get("标题") or "").strip()
                if not url or not title:
                    continue
                clean = clean_url(url, cfg)
                pid = extract_product_id(clean, cfg)
                if pid and pid in seen_ids:
                    continue
                if pid:
                    seen_ids.add(pid)
                # 只加从其他源没覆盖到的
                if pid and pid not in {i["商品ID"] for i in all_items}:
                    all_items.append({
                        "品类": fname.replace(".csv", ""),
                        "策略": "",
                        "排名": 0,
                        "品牌": row.get("brand", row.get("shop_name", "")),
                        "标题": title,
                        "价格": 0.0,
                        "型号": pid,
                        "商品ID": pid,
                        "详情页链接": clean,
                        "状态": "",
                    })

    all_items = all_items[:top_n]
    print(f"  ─────────────────────────────")
    print(f"  总计: {len(all_items)} 个商品")
    return all_items


# ═══════════════════════════════════════════════════════════════
#  主入口：根据平台调度
# ═══════════════════════════════════════════════════════════════

def build_download_list(
    cfg: PlatformConfig,
    top_n: int = 10,
    strategy_filter: str = "",
    category_filter: str = "",
) -> list[dict]:
    """根据平台配置构建下载清单。"""
    # 定位选品目录
    if cfg.default_listing_subdir:
        listing_dir = os.path.join(cfg.data_dir, cfg.default_listing_subdir)
    else:
        listing_dir = cfg.data_dir

    if not os.path.isdir(listing_dir):
        # 兜底：data_dir 本身
        listing_dir = cfg.data_dir

    if not os.path.isdir(listing_dir):
        raise FileNotFoundError(f"数据目录不存在: {listing_dir}")

    print(f"  平台:     {cfg.label}")
    print(f"  来源:     {listing_dir}")
    print(f"  Top:      {top_n}")
    print(f"  策略:     {strategy_filter or '(全部)'}")
    print()

    # 品类过滤
    if category_filter and cfg.default_listing_subdir:
        sub = os.path.join(listing_dir, category_filter)
        if not os.path.isdir(sub):
            print(f"  ⚠ 品类 [{category_filter}] 不存在，扫描全部")
        else:
            listing_dir = sub
            print(f"  品类:     {category_filter}")
            print()

    # 分平台扫描
    if cfg.name == "zkh":
        items = scan_zkh_listing(listing_dir, top_n, strategy_filter, cfg)
    elif cfg.name == "1688":
        items = scan_1688_listing(listing_dir, top_n, strategy_filter, cfg)
    else:
        items = scan_zkh_listing(listing_dir, top_n, strategy_filter, cfg)

    return items


# ═══════════════════════════════════════════════════════════════
#  导出 Excel
# ═══════════════════════════════════════════════════════════════

def export_to_excel(items: list[dict], output_path: str, platform_label: str):
    """导出为 Excel（可点击链接 + 策略颜色 + 筛选）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("❌ 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "详情页下载清单"

    headers = ["序号", "品类", "策略", "排名", "品牌", "标题",
               "价格", "型号", "商品ID", "详情页链接", "状态"]
    col_widths = [6, 10, 10, 6, 16, 50, 10, 30, 14, 50, 12]

    hfont = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    body_font = Font(name="微软雅黑", size=10)
    align_c = Alignment(horizontal="center", vertical="center")
    align_l = Alignment(vertical="center", wrap_text=True)

    strategy_colors = {
        "🔥 必上": "FFF2CC",
        "👍 推荐": "D9E2F3",
        "💡 暗马": "E2EFDA",
        "📌 关注": "FCE4D6",
        "🔝 代表品": "E2EFDA",
    }

    for idx, item in enumerate(items, 1):
        rn = idx + 1
        vals = [
            idx, item["品类"], item["策略"],
            item["排名"] if item["排名"] > 0 else "",
            item["品牌"], item["标题"],
            item["价格"] if item["价格"] > 0 else "",
            item["型号"], item["商品ID"],
            item["详情页链接"], item["状态"],
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=rn, column=ci)
            if ci == 10 and item["详情页链接"]:
                cell.value = item["详情页链接"]
                cell.hyperlink = item["详情页链接"]
                cell.font = link_font
            else:
                cell.value = val
                cell.font = body_font
            cell.alignment = align_c if ci in (1, 3, 4, 7, 9, 11) else align_l
            cell.border = border

        sc = strategy_colors.get(item["策略"])
        if sc:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=rn, column=ci).fill = PatternFill(
                    start_color=sc, end_color=sc, fill_type="solid"
                )

    # 自动筛选 + 冻结
    last_col = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(items) + 1}"
    ws.freeze_panes = "A2"

    # 说明 sheet
    ws2 = wb.create_sheet("使用说明", 0)
    lines = [
        f"\U0001f4cb {platform_label} 详情页下载清单 — 使用说明",
        "",
        "工作流:",
        "  1. 打开本 Excel，查看需要下载的商品",
        '  2. 点击\xe2\x80\x9c详情页链接\xe2\x80\x9d列的超链接 → 浏览器打开详情页',
        '  3. Ctrl+S → 选择\xe2\x80\x9c网页，仅HTML\xe2\x80\x9d → 保存',
        f"  4. 下载完成后，在\xe2\x80\x9c状态\xe2\x80\x9d列标记 \u2713",
        "  5. 运行批量解析",
        "",
        f"  \u2192 {platform_label} 详情目录建议:",
    ]
    if platform_label == "震坤行":
        lines.append("       data/ZKH/震坤行/details/")
        lines.append("     python -m platforms.zkh.collect_detail data/ZKH/震坤行/details/ --batch")
    else:
        lines.append("       data/1688/details/")
    lines.append("")
    lines.append("提示:")
    lines.append("  - 可以使用筛选功能按品类或策略分组下载")
    lines.append("  - 商品ID 即文件名，建议以此命名保存的 HTML")

    for i, line in enumerate(lines, 1):
        ws2.cell(row=i, column=1, value=line).font = Font(name="微软雅黑", size=11)
    ws2.column_dimensions["A"].width = 80

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"\n\U0001f4c1 已生成: {output_path}")
    print(f"   共 {len(items)} 个商品")


# ═══════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="选品清单 \u2192 详情页下载清单 (跨平台)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python -m selection.download_list --platform zkh --all --top 10
  python -m selection.download_list --platform 1688 --top 10
  python -m selection.download_list --platform zkh --category 室内灯具 --top 20
        """,
    )
    parser.add_argument("--platform", choices=list(PLATFORMS.keys()),
                        default="zkh", help="平台（默认 zkh）")
    parser.add_argument("--all", "-a", action="store_true",
                        help="处理所有品类")
    parser.add_argument("--category", "-c",
                        help="指定品类（ZKH: 目录名; 1688: 忽略）")
    parser.add_argument("--top", "-n", type=int, default=10,
                        help="取前 N 个（默认 10）")
    parser.add_argument("--strategy", "-s",
                        help='按策略筛选（如 \u300c必上\u300d）')
    parser.add_argument("--output", "-o", default="",
                        help="输出 Excel 路径")
    args = parser.parse_args()

    cfg = PLATFORMS[args.platform]

    # 生成输出路径
    if args.output:
        out_path = args.output
    else:
        os.makedirs(cfg.output_dir, exist_ok=True)
        parts = [f"下载清单_{cfg.label}"]
        if args.category:
            parts.append(args.category)
        if args.strategy:
            parts.append(args.strategy.replace(" ", ""))
        parts.append(f"top{args.top}")
        out_path = os.path.join(cfg.output_dir, "_".join(parts) + ".xlsx")

    print("\u2550" * 55)
    print(f"  选品清单 \u2192 详情页下载清单 ({cfg.label})")
    print("\u2550" * 55)

    items = build_download_list(
        cfg=cfg,
        top_n=args.top,
        strategy_filter=args.strategy or "",
        category_filter=args.category or "",
    )

    if not items:
        print("\n\u26a0 没有符合条件的商品")
        sys.exit(0)

    export_to_excel(items, out_path, cfg.label)


if __name__ == "__main__":
    main()
