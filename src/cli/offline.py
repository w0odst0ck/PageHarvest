"""
PageHarvest 离线解析 CLI

从 ZIP 文件解析电商商品数据（搜索页/详情页），
输出 CSV / TXT / XLSX / JSON 到指定目录。

用法:
    python -m cli.offline input.zip --mode search --out ./results/
    python -m cli.offline input.zip --mode detail --out ./results/ --verbose

产品策略: 离线分析 = 产品，在线采集 = 开源
"""

import argparse
import csv
import io
import json
import os
import re
import shutil
import sys
import tempfile
import traceback
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from io import BytesIO

# ── 项目内依赖 ──
try:
    from api.engine import process_upload
    from core.detail_parser import parse_detail, detect_platform
except ImportError:
    _base = Path(__file__).resolve().parent.parent.parent  # PageHarvest/
    _src = _base / "src"
    if str(_src) not in sys.path:
        sys.path.insert(0, str(_src))
    from api.engine import process_upload
    from core.detail_parser import parse_detail, detect_platform


# ═══════════════════════════════════════════════════════════════
#  工具：解压 ZIP
# ═══════════════════════════════════════════════════════════════

def extract_zip(zip_bytes: bytes) -> tuple[Path, list[Path]]:
    """解压到临时目录，返回 (tmp_dir, [文件路径列表])"""
    tmp = Path(tempfile.mkdtemp(prefix="ph_"))
    zpath = tmp / "upload.zip"
    zpath.write_bytes(zip_bytes)
    with zipfile.ZipFile(zpath, "r") as z:
        z.extractall(tmp)
    files = sorted(tmp.rglob("*")) if any(tmp.iterdir()) else []
    return tmp, [f for f in files if f.is_file() and not f.name.startswith(".")]


# ═══════════════════════════════════════════════════════════════
#  核心处理函数（返回 dict，供 Flask 和 CLI 共同使用）
# ═══════════════════════════════════════════════════════════════

def process_search(zip_bytes: bytes) -> dict:
    """搜索页解析 → 返回 display dict（与 Flask 期望的格式兼容）

    通过 api.engine.process_upload 路由。
    """
    pipeline_result = process_upload(zip_bytes)

    if not pipeline_result.success:
        return {"status": "error", "error": pipeline_result.error}

    display = {
        "status": "done",
        "mode": "search",
        "platform": pipeline_result.platform,
        "product_count": pipeline_result.product_count,
        "csv_content": pipeline_result.csv_content,
        "txt_content": pipeline_result.txt_content,
    }
    if pipeline_result.xlsx_bytes:
        display["has_xlsx"] = True
        display["_xlsx_bytes"] = pipeline_result.xlsx_bytes

    if pipeline_result.csv_content:
        reader = csv.DictReader(io.StringIO(pipeline_result.csv_content))
        display["table"] = list(reader)
        display["columns"] = reader.fieldnames or []

    return display


def process_detail(zip_bytes: bytes, images_output_dir: Path | None = None) -> dict:
    """详情页解析 → 返回 display dict（与 Flask 期望的格式兼容）

    直接调用 core.detail_parser，不经过 api.engine。
    images_output_dir: 主图复制目标目录（None = 不复制）
    """
    tmp_dir, files = extract_zip(zip_bytes)
    html_files = [f for f in files if f.suffix.lower() == ".html"]

    if not html_files:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return {"status": "error", "error": "ZIP 中未找到 HTML 文件"}

    results = []
    for i, html_path in enumerate(html_files):
        html = html_path.read_text(encoding="utf-8", errors="replace")
        platform = detect_platform(html)
        detail = parse_detail(html)

        if detail:
            # _files/ 目录补充图片（浏览器保存页面的静态资源目录）
            files_dir = html_path.parent / (html_path.stem + "_files")
            _supplement_images_from_files_dir(detail, files_dir, images_output_dir, i)

            # 相对路径 → 绝对路径
            _resolve_relative_image_paths(detail, html_path)

            results.append({
                "file": html_path.name,
                "platform": platform or "未知",
                "product_id": detail.product_id,
                "title": detail.title[:80] if detail.title else "",
                "brand": detail.brand or "-",
                "spec": detail.spec or "-",
                "price_min": detail.price_min,
                "price_max": detail.price_max,
                "attributes": len(detail.attributes or {}),
                "sku_count": detail.sku_count,
                "image_count": len(detail.main_images or []),
                "status": "✅",
                "_raw": detail,
            })
        else:
            results.append({
                "file": html_path.name,
                "platform": platform or "未知",
                "product_id": "",
                "title": "",
                "brand": "",
                "spec": "",
                "price_min": 0,
                "price_max": 0,
                "attributes": 0,
                "sku_count": 0,
                "image_count": 0,
                "status": "❌",
                "_raw": None,
            })

    # ── 构建 CSV ──
    csv_buf = io.StringIO()
    writer = csv.writer(csv_buf)
    writer.writerow(["文件", "平台", "商品ID", "标题", "品牌", "型号",
                     "最低价", "最高价", "属性数", "SKU数", "主图数", "状态"])
    for r in results:
        writer.writerow([r["file"], r["platform"], r["product_id"], r["title"],
                         r["brand"], r["spec"], r["price_min"], r["price_max"],
                         r["attributes"], r["sku_count"], r["image_count"], r["status"]])
    csv_content = csv_buf.getvalue()

    # ── 构建 TXT 报告 ──
    txt_content = _build_detail_txt(results)

    # ── 构建 XLSX ──
    ok = [r for r in results if r["status"] == "✅"]
    xlsx_bytes = _build_detail_xlsx(results)

    # 图片预览数据
    detail_images_preview = []
    detail_desc_images_preview = []
    for r in ok:
        raw = r.get("_raw")
        if raw:
            if raw.main_images:
                detail_images_preview.append({
                    "title": f"{raw.brand or '?'} - {(raw.title or '?')[:40]}",
                    "images": raw.main_images[:20],
                })
            if raw.detail_images:
                detail_desc_images_preview.append({
                    "title": f"{raw.brand or '?'} - {(raw.title or '?')[:40]}",
                    "images": raw.detail_images[:10],
                })

    # 完整解析 JSON
    import dataclasses
    raw_json = []
    for r in results:
        item = {k: v for k, v in r.items() if k != "_raw"}
        raw = r.get("_raw")
        if raw:
            item["_parsed"] = dataclasses.asdict(raw)
        else:
            item["_parsed"] = None
        raw_json.append(item)

    display = {
        "status": "done",
        "mode": "detail",
        "product_count": len(results),
        "success_count": len(ok),
        "csv_content": csv_content,
        "txt_content": txt_content,
        "table": [{k: v for k, v in r.items() if k != "_raw"} for r in results],
        "columns": ["文件", "平台", "商品ID", "标题", "品牌", "型号",
                    "最低价", "最高价", "属性数", "SKU数", "主图数", "状态"],
        "_detail_raw": raw_json,
        "_detail_images": detail_images_preview,
        "_detail_desc_images": detail_desc_images_preview,
    }
    if xlsx_bytes:
        display["has_xlsx"] = True
        display["_xlsx_bytes"] = xlsx_bytes

    # 清理临时目录
    shutil.rmtree(tmp_dir, ignore_errors=True)

    return display


def _supplement_images_from_files_dir(
    detail, files_dir: Path, images_output_dir: Path | None, product_idx: int
):
    """从浏览器 _files/ 目录补充主图"""
    if not files_dir.is_dir() or len(detail.main_images) >= 5:
        return
    candidates = []
    for img_file in files_dir.iterdir():
        if img_file.suffix.lower() not in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".avif"):
            continue
        if img_file.stat().st_size < 10240:
            continue
        candidates.append(img_file)
    candidates.sort(key=lambda f: f.stat().st_size, reverse=True)
    selected = candidates[:20]

    if not selected:
        return

    if images_output_dir:
        # CLI 模式：复制到 images/{idx}_{filename}
        per_product = images_output_dir
        per_product.mkdir(parents=True, exist_ok=True)
        img_urls = []
        for j, img_file in enumerate(selected):
            try:
                ext = img_file.suffix or ".jpg"
                dst_name = f"{product_idx:04d}_{j:02d}{ext}"
                shutil.copy2(str(img_file), str(per_product / dst_name))
                img_urls.append(str(per_product / dst_name))
            except Exception:
                pass
        if img_urls:
            detail.main_images = img_urls
    else:
        # Flask 模式：直接记录路径（后续由 Flask 路由处理）
        img_urls = [str(f) for f in selected]
        detail.main_images = img_urls


def _resolve_relative_image_paths(detail, html_path: Path):
    """解析器提取到的相对路径 → 绝对路径"""
    if not detail.main_images:
        return
    resolved = []
    for img_path in detail.main_images:
        if isinstance(img_path, str) and img_path.startswith("./"):
            abs_path = html_path.parent / img_path[2:]
            if abs_path.exists():
                resolved.append(str(abs_path))
            else:
                resolved.append(img_path)
        else:
            resolved.append(img_path)
    detail.main_images = resolved


def _build_detail_txt(results: list[dict]) -> str:
    """构建详情页 TXT 报告"""
    ok = [r for r in results if r.get("status") == "✅"]
    fail = [r for r in results if r.get("status") == "❌"]
    lines = ["PageHarvest 详情页解析报告", ""]
    platforms = set(r.get("platform", "未知") for r in ok)
    lines.append(f"共 {len(results)} 页，成功 {len(ok)}，失败 {len(fail)}")
    lines.append(f"涉及平台: {', '.join(sorted(platforms))}")
    lines.append("")
    for r in ok:
        lines.append(f"  ✅ {r.get('brand', '-'):12s} ¥{r.get('price_min', 0):<8.2f}  "
                     f"{r.get('sku_count', 0)} SKU  {r.get('attributes', 0)} 属性  "
                     f"{r.get('image_count', 0)} 图  {str(r.get('title', ''))[:35]}")
    return "\n".join(lines)


def _build_detail_xlsx(results: list[dict]) -> bytes:
    """构建详情页 Excel 报告"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return b""

    import dataclasses

    wb = Workbook()
    ws = wb.active
    ws.title = "解析结果"

    all_cols = ["file", "platform", "product_id", "title", "brand", "spec",
                "price_min", "price_max", "attributes_count", "sku_count",
                "image_count", "status"]
    parsed_cols = ["ship_from", "sales_count", "yearly_sales", "repurchase_rate",
                   "listing_date", "product_code", "min_order"]
    all_cols += [f"parsed.{c}" for c in parsed_cols]
    all_cols += ["attributes", "sku_matrix", "main_images", "detail_images"]

    ws.append(all_cols)
    _style_header_row(ws, all_cols)

    for r in results:
        raw = r.get("_raw")
        parsed = dataclasses.asdict(raw) if raw else {}
        row = []
        for c in ["file", "platform", "product_id", "title", "brand", "spec",
                  "price_min", "price_max"]:
            row.append(r.get(c, ""))
        row.append(len(parsed.get("attributes", {})))
        row.append(r.get("sku_count", 0))
        row.append(r.get("image_count", 0))
        row.append(r.get("status", ""))
        for c in parsed_cols:
            row.append(parsed.get(c, ""))
        row.append(str(parsed.get("attributes", {})))
        row.append(str(parsed.get("sku_matrix", [])))
        row.append("\n".join(parsed.get("main_images", [])))
        row.append("\n".join(parsed.get("detail_images", [])))
        ws.append(row)

    # 各商品独立 Sheet
    for r in results:
        raw = r.get("_raw")
        if not raw:
            continue
        parsed = dataclasses.asdict(raw)
        safe_name = re.sub(r'[\\/*?\[\]:]', '_',
                           f"{parsed.get('title', '')[:30] or parsed.get('product_id', '')}")
        if not safe_name:
            safe_name = f"item_{results.index(r)}"
        sheet = wb.create_sheet(title=safe_name[:31])

        info_rows = [
            ("商品ID", parsed.get("product_id", "")),
            ("标题", parsed.get("title", "")),
            ("品牌", parsed.get("brand", "")),
            ("型号", parsed.get("spec", "")),
            ("最低价", parsed.get("price_min", 0)),
            ("最高价", parsed.get("price_max", 0)),
        ]
        for kv in info_rows:
            sheet.append(kv)

        attrs = parsed.get("attributes", {})
        if attrs:
            sheet.append([])
            sheet.append(["属性", "值"])
            for k, v in attrs.items():
                sheet.append([k, v])

        sku_matrix = parsed.get("sku_matrix", [])
        if sku_matrix:
            sheet.append([])
            sheet.append(["SKU 规格", "价格"])
            for sku in sku_matrix:
                if isinstance(sku, dict):
                    sheet.append([sku.get("spec", ""), sku.get("price", "")])

        main_imgs = parsed.get("main_images", [])
        if main_imgs:
            sheet.append([])
            sheet.append(["主图链接"])
            for u in main_imgs[:20]:
                sheet.append([u])

    _auto_width(ws)

    xlsx_buf = BytesIO()
    wb.save(xlsx_buf)
    xlsx_buf.seek(0)
    return xlsx_buf.read()


def _style_header_row(ws, headers: list[str]):
    """给首行加粗 + 蓝色背景"""
    try:
        from openpyxl.styles import Font, PatternFill
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    except ImportError:
        pass


def _auto_width(ws):
    """自动列宽"""
    for col in ws.columns:
        try:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════
#  写文件辅助（CLI 模式）
# ═══════════════════════════════════════════════════════════════

def _write_csv(out_dir: Path, csv_content: str, filename: str = "report.csv"):
    if not csv_content:
        return None
    path = out_dir / filename
    path.write_text(csv_content, encoding="utf-8-sig")
    return path


def _write_txt(out_dir: Path, txt_content: str, filename: str = "report.txt"):
    if not txt_content:
        return None
    path = out_dir / filename
    path.write_text(txt_content, encoding="utf-8")
    return path


def _write_xlsx(out_dir: Path, xlsx_bytes: bytes, filename: str = "report.xlsx"):
    if not xlsx_bytes:
        return None
    path = out_dir / filename
    path.write_bytes(xlsx_bytes)
    return path


def _write_json(out_dir: Path, data: dict, filename: str = "report.json"):
    path = out_dir / filename
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path


# ═══════════════════════════════════════════════════════════════
#  离线解析入口（CLI 模式：解析 + 写文件）
# ═══════════════════════════════════════════════════════════════

def run_offline_search(zip_bytes: bytes, out_dir: Path, verbose: bool = False) -> int:
    """搜索页离线解析 → 写入 out_dir"""
    if verbose:
        print("  📦 正在解压...")

    display = process_search(zip_bytes)

    if display.get("status") == "error":
        print(f"  ❌ 解析失败: {display.get('error', '')}")
        return 1

    if verbose:
        print(f"  🔍 平台: {display.get('platform', '')}  |  商品数: {display.get('product_count', 0)}")

    # 构建完整 JSON（含原始数据）
    report_data = {
        "mode": "search",
        "platform": display.get("platform"),
        "product_count": display.get("product_count"),
        "csv": display.get("csv_content"),
        "txt": display.get("txt_content"),
        "has_xlsx": display.get("has_xlsx", False),
    }

    # 写文件
    _write_csv(out_dir, display.get("csv_content", ""))
    _write_txt(out_dir, display.get("txt_content", ""))

    xlsx_bytes = display.get("_xlsx_bytes")
    if xlsx_bytes:
        _write_xlsx(out_dir, xlsx_bytes)

    _write_json(out_dir, report_data)

    if verbose:
        for f in out_dir.iterdir():
            if f.is_file():
                print(f"     📄 {f.name}")

    print(f"  ✅ 搜索页解析完成，共 {display.get('product_count', 0)} 个商品")
    return 0


def run_offline_detail(zip_bytes: bytes, out_dir: Path, verbose: bool = False) -> int:
    """详情页离线解析 → 写入 out_dir"""
    if verbose:
        print("  📦 正在解压 ZIP...")

    images_dir = out_dir / "images"
    display = process_detail(zip_bytes, images_output_dir=images_dir)

    if display.get("status") == "error":
        print(f"  ❌ 解析失败: {display.get('error', '')}")
        return 1

    # 写文件
    _write_csv(out_dir, display.get("csv_content", ""))
    _write_txt(out_dir, display.get("txt_content", ""))

    xlsx_bytes = display.get("_xlsx_bytes")
    if xlsx_bytes:
        _write_xlsx(out_dir, xlsx_bytes)

    # 构建完整 JSON
    report_data = {
        "mode": "detail",
        "product_count": display.get("product_count", 0),
        "success_count": display.get("success_count", 0),
        "csv": display.get("csv_content"),
        "txt": display.get("txt_content"),
        "has_xlsx": bool(xlsx_bytes),
        "detail_raw": display.get("_detail_raw", []),
    }
    _write_json(out_dir, report_data)

    if verbose:
        print(f"  📄 输出文件:")
        for f in out_dir.iterdir():
            if f.is_file():
                print(f"     {f.name}")
        img_count = len(list(images_dir.glob("*"))) if images_dir.exists() else 0
        if img_count:
            print(f"  🖼️ 主图: {img_count} 张 ({images_dir})")

    ok = display.get("success_count", 0)
    total = display.get("product_count", 0)
    print(f"  ✅ 详情页解析完成: {ok}/{total} 页成功")
    return 0


def run_offline(zip_bytes: bytes, out_dir: Path, mode: str, verbose: bool = False) -> int:
    """统一入口：根据 mode 路由到 search 或 detail 管道"""
    if mode == "search":
        return run_offline_search(zip_bytes, out_dir, verbose)
    elif mode == "detail":
        return run_offline_detail(zip_bytes, out_dir, verbose)
    else:
        print(f"❌ 不支持的 mode: {mode}（可选: search / detail）")
        return 1


# ═══════════════════════════════════════════════════════════════
#  CLI 入口
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="PageHarvest 离线解析 — 从 ZIP 文件提取电商商品数据",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python -m cli.offline input.zip --mode search
  python -m cli.offline input.zip --mode detail --out ./results/ --verbose
        """,
    )
    parser.add_argument("input_zip", help="待解析的 ZIP 文件路径")
    parser.add_argument("--mode", required=True, choices=["search", "detail"],
                        help="解析模式: search（搜索页选品）/ detail（详情页解析）")
    parser.add_argument("--out", default="./results/",
                        help="输出目录（默认 ./results/）")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="输出详细进度信息")

    args = parser.parse_args()

    zip_path = Path(args.input_zip)
    if not zip_path.exists():
        print(f"❌ 文件不存在: {zip_path}")
        sys.exit(1)

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.verbose:
        print(f"✦ PageHarvest 离线解析")
        print(f"   输入: {zip_path}")
        print(f"   模式: {args.mode}")
        print(f"   输出: {out_dir.resolve()}")

    try:
        zip_bytes = zip_path.read_bytes()
        exit_code = run_offline(zip_bytes, out_dir, args.mode, args.verbose)
        sys.exit(exit_code)
    except Exception as e:
        print(f"❌ 解析异常: {e}")
        if args.verbose:
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
