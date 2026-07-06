#!/usr/bin/env python3
"""
详情页批量下载器 (Batch Detail Downloader)
============================================

从下载清单 Excel/CSV 中读取商品 URL，批量下载详情页 HTML。
自动处理各平台 WAF 挑战（通过 Playwright 浏览器渲染）。

用法:
    # 从下载清单 Excel 批量下载
    python -m pipeline.detail_downloader data/ZKH/震坤行/download_list/下载清单_震坤行_all_top10.xlsx

    # 从 CSV 下载
    python -m pipeline.detail_downloader data/1688/top_products_urls.csv

    # 指定输出目录和并发数
    python -m pipeline.detail_downloader list.xlsx --output data/ZKH/震坤行/details/ --workers 3

安装:
    pip install playwright
    python -m playwright install chromium
"""

import os
import re
import sys
import csv
import json
import time
import argparse
import logging
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

_THIS_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = _THIS_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

logger = logging.getLogger("detail_downloader")


# ── 平台判断 ──
def detect_platform(url: str) -> str:
    """从 URL 判断平台"""
    if "zkh.com" in url or "震坤行" in url:
        return "zkh"
    if "1688.com" in url or "detail.1688" in url:
        return "1688"
    if "jd.com" in url:
        return "jd"
    return "unknown"


def detail_url_1688(product_id: str) -> str:
    return f"https://detail.1688.com/offer/{product_id}.html"


def detail_url_zkh(product_id: str) -> str:
    return f"https://www.zkh.com/product/detail/{product_id}.html"


# ── 读取下载清单 ──
def read_download_list(path: str) -> list[dict]:
    """从 Excel 或 CSV 读取下载清单"""
    items = []

    if path.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            print("❌ 需要 openpyxl: pip install openpyxl")
            sys.exit(1)
        wb = openpyxl.load_workbook(path)
        ws = wb["详情页下载清单"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
            url = (row.get("详情页链接") or "").strip()
            pid = (row.get("商品ID") or "").strip()
            if url or pid:
                items.append(row)
    else:
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                url = (row.get("detail_url") or row.get("详情页链接") or "").strip()
                pid = (row.get("offer_id") or row.get("商品ID") or "").strip()
                if url or pid:
                    items.append(row)

    return items


# ── Playwright 下载器 ──
_browser = None
_page = None


def _ensure_browser():
    """全局单例浏览器（避免反复启动）"""
    global _browser, _page
    if _browser is not None:
        return _page

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print()
        print("=" * 55)
        print("  需要安装 Playwright:")
        print("    pip install playwright")
        print("    python -m playwright install chromium")
        print("=" * 55)
        raise

    p = sync_playwright().start()
    _browser = p.chromium.launch(
        headless=True,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox",
        ],
    )
    _page = _browser.new_page()
    # 反检测
    _page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
    """)
    return _page


def download_page(product_id: str, platform: str, url: str = "") -> str | None:
    """用 Playwright 下载一个详情页，返回 HTML（失败返回 None）"""
    page = _ensure_browser()
    target_url = url or (
        detail_url_1688(product_id) if platform == "1688"
        else detail_url_zkh(product_id)
    )

    try:
        # 导航
        logger.debug("  GET %s", target_url)
        page.goto(target_url, wait_until="domcontentloaded", timeout=30000)

        # 等待 WAF 挑战解决（如果有）和页面渲染
        # 1688: 等待 offer_name / price 出现
        # ZKH:  等待 params-wrap 或 sku-number 出现
        wait_selectors = {
            "1688": ["#offer-name", ".offer-name", '[class*="title"]', '[class*="price"]'],
            "zkh":  [".params-wrap", ".sku-number", ".gallery-wrap", '[class*="product"]'],
        }
        selectors = wait_selectors.get(platform, [])
        for sel in selectors:
            try:
                page.wait_for_selector(sel, timeout=15000)
                break
            except:
                continue

        # 额外等待页面稳定
        time.sleep(2)

        html = page.content()
        if not html or len(html) < 5000:
            logger.warning("  ⚠ HTML 过短 (%d bytes), 可能被拦截", len(html))
            return None

        return html

    except Exception as e:
        logger.warning("  ⚠ 下载失败: %s", e)
        return None


def close_browser():
    """关闭浏览器"""
    global _browser, _page
    if _browser:
        _browser.close()
        _browser = None
        _page = None


# ── 保存 HTML ──
def save_html(html: str, output_dir: str, product_id: str, platform: str) -> str:
    """保存 HTML 到指定目录"""
    os.makedirs(output_dir, exist_ok=True)
    fname = f"{product_id}.html"
    fpath = os.path.join(output_dir, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        f.write(html)
    return fpath


# ── 单任务 ──
def download_one(item: dict, output_dir: str) -> dict:
    """下载一个商品，返回结果"""
    url = (item.get("详情页链接") or item.get("detail_url") or "").strip()
    pid = (item.get("商品ID") or item.get("offer_id") or "").strip()
    platform = detect_platform(url) if url else "zkh"

    if not pid:
        # 从 URL 提取
        m = re.search(r'/item/([^./?]+)', url) if "zkh" in url else re.search(r'/offer/(\d+)', url)
        pid = m.group(1) if m else ""

    if not pid:
        return {"id": "?", "status": "FAIL", "reason": "无商品ID"}

    # 目标目录
    plat_dir = os.path.join(output_dir, "ZKH" if platform == "zkh" else "1688", "details")
    os.makedirs(plat_dir, exist_ok=True)

    # 已下载则跳过
    existing = os.path.join(plat_dir, f"{pid}.html")
    if os.path.exists(existing) and os.path.getsize(existing) > 5000:
        return {"id": pid, "status": "SKIP", "file": existing}

    html = download_page(pid, platform, url)
    if html is None:
        return {"id": pid, "status": "FAIL", "reason": "下载失败"}

    fpath = save_html(html, plat_dir, pid, platform)
    size_kb = len(html) // 1024
    return {"id": pid, "status": "OK", "file": fpath, "size": size_kb}


def main():
    parser = argparse.ArgumentParser(
        description="详情页批量下载器 — 尝试自动下载（实验性，默认走人工）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python -m pipeline.detail_downloader 下载清单.xlsx
  python -m pipeline.detail_downloader top_products_urls.csv
  python -m pipeline.detail_downloader list.xlsx --workers 3 --output data/ZKH/震坤行/details/
        """,
    )
    parser.add_argument("input", help="下载清单 Excel 或 CSV")
    parser.add_argument("--output", "-o", default=str(PROJECT_ROOT / "data"),
                        help="输出根目录（默认 data/）")
    parser.add_argument("--workers", "-w", type=int, default=2,
                        help="并发数（默认 2，太大易触发反爬）")
    parser.add_argument("--limit", "-n", type=int, default=0,
                        help="限制下载数量（0=全部）")
    parser.add_argument("--resume", "-r", action="store_true",
                        help="跳过已下载的文件")
    parser.add_argument("--verbose", "-v", action="store_true", help="详细日志")
    args = parser.parse_args()

    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(level=level, format="%(message)s")

    # ── 读取清单 ──
    if not os.path.exists(args.input):
        print(f"❌ 文件不存在: {args.input}")
        sys.exit(1)

    print("═" * 55)
    print("  详情页批量下载器 (实验性)")
    print("═" * 55)
    print("  ⚠ 自动化下载依赖 Playwright + 登录态")
    print("  ⚠ 如果遇到 WAF 滑块验证码，请切换至人工下载")
    print()

    items = read_download_list(args.input)
    if not items:
        print("❌ 清单中无有效商品")
        sys.exit(1)

    total = len(items)
    if args.limit and args.limit < total:
        items = items[:args.limit]

    print(f"  清单:     {args.input}")
    print(f"  商品:     {len(items)}/{total}")
    print(f"  并发:     {args.workers}")
    print(f"  输出:     {args.output}")
    print()

    # ── 初始化浏览器 ──
    try:
        _ensure_browser()
    except Exception as e:
        print(f"\n❌ 浏览器初始化失败: {e}")
        sys.exit(1)

    # ── 批量下载 ──
    results = {"OK": 0, "SKIP": 0, "FAIL": 0}
    failures = []

    if args.workers > 1 and args.workers <= 5:
        # 并发
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            futures = {pool.submit(download_one, item, args.output): item for item in items}
            for i, future in enumerate(as_completed(futures), 1):
                r = future.result()
                results[r["status"]] = results.get(r["status"], 0) + 1
                _print_progress(i, len(items), r)
                if r["status"] == "FAIL":
                    failures.append(r)
    else:
        # 串行（默认）
        for i, item in enumerate(items, 1):
            r = download_one(item, args.output)
            results[r["status"]] = results.get(r["status"], 0) + 1
            _print_progress(i, len(items), r)
            if r["status"] == "FAIL":
                failures.append(r)

    close_browser()

    # ── 报告 ──
    print()
    print("═" * 55)
    print(f"  完成: {results.get('OK', 0)} OK  /  {results.get('SKIP', 0)} 跳过  /  {results.get('FAIL', 0)} 失败")
    if failures:
        print()
        print("  失败明细:")
        for f in failures:
            print(f"    ❌ {f['id']:>14}: {f.get('reason', '?')}")
    print("═" * 55)


def _print_progress(current: int, total: int, result: dict):
    pid = result["id"]
    status = result["status"]
    if status == "OK":
        size = result.get("size", 0)
        print(f"  [{current:>3}/{total}] ✅ {pid:>14}  ({size} KB)")
    elif status == "SKIP":
        print(f"  [{current:>3}/{total}] ⏭ {pid:>14}  已存在")
    else:
        reason = result.get("reason", "")
        print(f"  [{current:>3}/{total}] ❌ {pid:>14}  {reason}")


if __name__ == "__main__":
    main()
